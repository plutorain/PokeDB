#import numpy as np
#import pandas as pd
from pandas import read_excel
from pandas import ExcelWriter
from pandas import DataFrame
from pandas import concat
import os
import enum

#QT GUI
import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import QObject, QThread, pyqtSignal, pyqtSlot

class JP_INDEX(enum.Enum):
    CARD_NO = 0
    CARD_TYPE = 9
    FROM = 11
    CARD_NAME = 13
    HP        = 15
    POKE_TYPE = 17
    SKILL_COST = 29
    SKILL_NAME = 26
    SKILL_TEXT = 32
    
    POKE_NO = 34
    SKILL_DAMAGE = 30
    
    CATEGORY = 36
    HEIGHT = 38
    WEIGHT = 40

    WEAK_TYPE = 19
    WEAK_MULTI = 20
    REGI_TYPE = 22
    REGI_MULTI = 23
    RUN_COUNT = 24  
    FLAVOR = 41
    ILLUSTRATOR  = 3
    RARITY = 2
    SERIES_MARK = 4
    REGULATION_MARK = 5

class KR_INDEX(enum.Enum):
    WR_COLLECT_NO=3 #1
    WR_POKE_NO=4 #35
    WR_RARITY=5 #3
    WR_CARDTYPE_NUM=6 #신규 1 Pokecard 2 Trainers 3 Energy
    WR_CARDTYPE=7 #37
    WR_NAME=8 #14
    WR_MON_TYPE=9 #18
    WR_HEIGHT=10 #39
    WR_WEIGHT=11 #41
    WR_DOX_INFO=12 #42
    WR_PRE_MON=13 #12

    WR_LEVEL=14

    WR_ILLUSTRATOR=15 #4
    WR_HP=16 #16
    WR_CARDMONTYPE=17 #18
    WR_WEAKNESS=18 #20
    WR_WEAKNESS_LVL=19 #21
    WR_RESISTANCE=20 # 23
    WR_RESISTANCE_LVL=21 #24
    WR_RETREAT=22 #25
    WR_TECHERG1=23 #30-1
    WR_TECHERG2=24 #30-2
    WR_TECHERG3=25 #30-3
    WR_TECHERG4=26 #30-4 #{G}풀,{R}불꽃,{W}물,{L}번개,{P}초,{F}격투,{D}악,{M}강철,{S}페어리,{C}무색,{Y}드래곤

    WR_DAMAGE1=27 #31-1
    WR_DAMAGE2=28 #31-2
    WR_DAMAGE3=29 #31-3
    WR_DAMAGE4=30 #31-4

    WR_AB_LABEL1=31 #27-1
    WR_AB_LABEL2=32 #27-2
    WR_AB_LABEL3=33 #27-3
    WR_AB_LABEL4=34 #27-4

    WR_ABILITY1=35 #33-1
    WR_ABILITY2=36 #33-2
    WR_ABILITY3=37 #33-3
    WR_ABILITY4=38 #33-4

    WR_POWERBODY=39
    WR_LEVEL_RULE=40

    WR_SERIES_MARK = 46
    WR_REGULATION_MARK = 49



class JP_DATA:
    def __init__(self, f_name):
        print("JP_DATA INIT")
        #self.xlsx = pd.read_excel(f_name, encoding='utf-8', keep_default_na=False) #read dataframe
        self.xlsx = read_excel(f_name, encoding='utf-8', keep_default_na=False) #read dataframe
        self.key = self.xlsx.keys()
        self.START_ROW = 0
        self.isLoadOK = False

        EOF = len(self.xlsx.index)
        print("JPDATA INDEX SIZE :", len(self.xlsx.index))

        while(1): #앞 세글자가 숫자일때까지 (1번 CARD위치찾기)
            cell = self.XLS_DATA(0 , self.START_ROW)
            if( (str(type(cell)) == "<class 'str'>") and cell[0:3].isnumeric() ):
                print ("find Card No.", cell[0:3])
                break
            self.START_ROW +=1
            if(self.START_ROW == EOF):
                print("JPDATA return None!!!")
                return None

        self.NOW_ROW = self.START_ROW
        self.isLoadOK = True
        #print(self.xlsx.index.stop)
    
    def XLS_DATA( self , col , row ):
        return self.xlsx[self.key[col]][row]
    
    def SetStartPos(self):
        self.NOW_ROW = self.START_ROW
        
    def Get_Card_Data(self):
        if(self.NOW_ROW >= self.xlsx.index.stop): #End of file
            return "EOF"
        
        tmp = POKE_CARD()
        depth = 0
        #Get Current Card No
        
        tmp.CARD_NO = self.XLS_DATA(JP_INDEX.CARD_NO.value,self.NOW_ROW)
       
        while(self.XLS_DATA(JP_INDEX.CARD_NO.value, self.NOW_ROW + depth) == tmp.CARD_NO[0]):
            depth += 1
            if(self.NOW_ROW + depth == self.xlsx.index.stop):
                break
            
        #print("depth : %d , NOW_ROW: %d" %(depth, self.NOW_ROW))
        
        tmp.CARD_TYPE = self.XLS_DATA(JP_INDEX.CARD_TYPE.value,self.NOW_ROW)
        tmp.CARD_NAME = self.XLS_DATA(JP_INDEX.CARD_NAME.value,self.NOW_ROW)
        tmp.HP = self.XLS_DATA(JP_INDEX.HP.value,self.NOW_ROW)
        tmp.POKE_TYPE = self.XLS_DATA(JP_INDEX.POKE_TYPE.value,self.NOW_ROW)
        
        tmp.WEAK_TYPE = self.XLS_DATA(JP_INDEX.WEAK_TYPE.value,self.NOW_ROW)
        tmp.WEAK_MULTI = self.XLS_DATA(JP_INDEX.WEAK_MULTI.value,self.NOW_ROW)
        tmp.REGI_TYPE = self.XLS_DATA(JP_INDEX.REGI_TYPE.value,self.NOW_ROW)
        tmp.REGI_MULTI = self.XLS_DATA(JP_INDEX.REGI_MULTI.value,self.NOW_ROW)
        
        tmp.RUN_COUNT = self.XLS_DATA(JP_INDEX.RUN_COUNT.value,self.NOW_ROW)
        tmp.FLAVOR = self.XLS_DATA(JP_INDEX.FLAVOR.value,self.NOW_ROW)
        tmp.ILLUSTRATOR = self.XLS_DATA(JP_INDEX.ILLUSTRATOR.value,self.NOW_ROW)
        
        tmp.FROM = self.XLS_DATA(JP_INDEX.FROM.value,self.NOW_ROW)
        tmp.POKE_NO = self.XLS_DATA(JP_INDEX.POKE_NO.value,self.NOW_ROW)
        tmp.CATEGORY = self.XLS_DATA(JP_INDEX.CATEGORY.value,self.NOW_ROW)
        tmp.HEIGHT = self.XLS_DATA(JP_INDEX.HEIGHT.value,self.NOW_ROW)
        tmp.WEIGHT = self.XLS_DATA(JP_INDEX.WEIGHT.value,self.NOW_ROW)
        tmp.RARITY = self.XLS_DATA(JP_INDEX.RARITY.value,self.NOW_ROW)
               
        for i in range(depth):
            tmp.SKILL_COST = self.XLS_DATA(JP_INDEX.SKILL_COST.value, self.NOW_ROW + i)        
            tmp.SKILL_NAME = self.XLS_DATA(JP_INDEX.SKILL_NAME.value, self.NOW_ROW + i)
            tmp.SKILL_TEXT = self.XLS_DATA(JP_INDEX.SKILL_TEXT.value, self.NOW_ROW + i)
            tmp.SKILL_DAMAGE = self.XLS_DATA(JP_INDEX.SKILL_DAMAGE.value, self.NOW_ROW + i)
        
        tmp.SERIES_MARK = self.XLS_DATA(JP_INDEX.SERIES_MARK.value, self.NOW_ROW)
        tmp.REGULATION_MARK = self.XLS_DATA(JP_INDEX.REGULATION_MARK.value, self.NOW_ROW)
        
        self.NOW_ROW += depth #다음 Card Data로 이동
        return tmp
        
    def export_txt_files(self, filename):
        f = open(filename, 'w', -1 ,"utf-8")
        while( self.NOW_ROW < self.xlsx.index.stop ):
            print("NOW_CARD : %s" % self.XLS_DATA(JP_INDEX.CARD_NO.value,self.NOW_ROW))
            f.write(self.Get_Card_Data().__str__())
        f.close()
        
class POKE_CARD:
    
    def __init__(self):
        #print("POKE_CARD_INIT")
        self._CARD_NO = []
        self._CARD_TYPE = []
        self._CARD_NAME = []
        self._HP       = []
        self._POKE_TYPE = []
        self._SKILL_COST = []
        self._SKILL_NAME = []
        self._SKILL_TEXT = []
        self._WEAK_TYPE = []
        
        self._WEAK_MULTI = []
        self._REGI_TYPE = []
        self._RUN_COUNT = []
        self._FLAVOR = []
        self._ILLUSTRATOR  = []
        self._SKILL_DAMAGE = []
        self._FROM = []
        self._POKE_NO = []
        self._CATEGORY = []
        self._HEIGHT = []
        self._WEIGHT = []
        self._RARITY = []
        self._REGI_MULTI = []

        self._SERIES_MARK = []
        self._REGULATION_MARK = []

    
    def _get_CARD_NO(self):
        return self._CARD_NO
    def _get_CARD_TYPE(self):
        return self._CARD_TYPE
    def _get_CARD_NAME (self):
        return self._CARD_NAME 
    def _get_HP(self):
        return self._HP
    def _get_POKE_TYPE (self):
        return self._POKE_TYPE 
    def _get_SKILL_COST (self):
        return self._SKILL_COST 
    def _get_SKILL_NAME (self):
        return self._SKILL_NAME 
    def _get_SKILL_TEXT (self):
        return self._SKILL_TEXT 
    def _get_WEAK_TYPE (self):
        return self._WEAK_TYPE 
    def _get_WEAK_MULTI(self):
        return self._WEAK_MULTI
    def _get_REGI_TYPE (self):
        return self._REGI_TYPE 
    def _get_RUN_COUNT  (self):
        return self._RUN_COUNT  
    def _get_FLAVOR(self):
        return self._FLAVOR
    def _get_ILLUSTRATOR(self):
        return self._ILLUSTRATOR
    def _get_SKILL_DAMAGE(self):
        return self._SKILL_DAMAGE
    


    def _set_CARD_NO (self , _CARD_NO = None):
        self._CARD_NO.append(_CARD_NO)
    def _set_CARD_TYPE (self , _CARD_TYPE = None):
        self._CARD_TYPE.append(_CARD_TYPE)
    def _set_CARD_NAME  (self , _CARD_NAME  = None):
        self._CARD_NAME.append(_CARD_NAME )
    def _set_HP (self , _HP = None):
        self._HP.append(_HP)
    def _set_POKE_TYPE  (self , _POKE_TYPE  = None):
        self._POKE_TYPE.append(_POKE_TYPE )
    def _set_SKILL_COST  (self , _SKILL_COST  = None):
        self._SKILL_COST.append(_SKILL_COST )
    def _set_SKILL_NAME  (self , _SKILL_NAME  = None):
        self._SKILL_NAME.append(_SKILL_NAME )
    def _set_SKILL_TEXT  (self , _SKILL_TEXT  = None):
        self._SKILL_TEXT.append(_SKILL_TEXT )
    def _set_WEAK_TYPE  (self , _WEAK_TYPE  = None):
        self._WEAK_TYPE.append(_WEAK_TYPE )
    def _set_WEAK_MULTI (self , _WEAK_MULTI = None):
        self._WEAK_MULTI.append(_WEAK_MULTI)
    def _set_REGI_TYPE  (self , _REGI_TYPE  = None):
        self._REGI_TYPE.append(_REGI_TYPE )
    def _set_RUN_COUNT   (self , _RUN_COUNT   = None):
        self._RUN_COUNT.append(_RUN_COUNT  )
    def _set_FLAVOR (self , _FLAVOR = None):
        self._FLAVOR.append(_FLAVOR)
    def _set_ILLUSTRATOR (self, _ILLUSTRATOR = None):
        self._ILLUSTRATOR.append(_ILLUSTRATOR)
    def _set_SKILL_DAMAGE (self , _SKILL_DAMAGE = None):
        self._SKILL_DAMAGE.append(_SKILL_DAMAGE)
    
    
    
    def _set_FROM (self , _FROM = None):
        self._FROM.append(_FROM)
    def _get_FROM(self):
        return self._FROM
        
    def _set_POKE_NO (self , _POKE_NO = None):
        self._POKE_NO.append(_POKE_NO)										
    def _get_POKE_NO(self):
        return self._POKE_NO

        						
    def _set_CATEGORY (self , _CATEGORY = None):
        self._CATEGORY.append(_CATEGORY)
    def _get_CATEGORY(self):
        return self._CATEGORY
        
    					
    def _set_HEIGHT (self , _HEIGHT = None):
        self._HEIGHT.append(_HEIGHT)										
    def _get_HEIGHT(self):
        return self._HEIGHT
        
    def _set_WEIGHT (self , _WEIGHT = None):
        self._WEIGHT.append(_WEIGHT)
    def _get_WEIGHT(self):
        return self._WEIGHT
        
    
    def _set_RARITY (self , _RARITY = None):
        self._RARITY.append(_RARITY)
    def _get_RARITY(self):
        return self._RARITY
   
    def _set_REGI_MULTI (self , _REGI_MULTI = None):
        self._REGI_MULTI.append(_REGI_MULTI)
    def _get_REGI_MULTI(self):
        return self._REGI_MULTI

    def _set_SERIES_MARK (self , _SERIES_MARK = None):
        self._SERIES_MARK.append(_SERIES_MARK)
    def _set_REGULATION_MARK (self , _REGULATION_MARK = None):
        self._REGULATION_MARK.append(_REGULATION_MARK)
    def _get_SERIES_MARK(self):
        return self._SERIES_MARK
    def _get_REGULATION_MARK(self):
        return self._REGULATION_MARK

    CARD_NO = property(_get_CARD_NO, _set_CARD_NO)
    CARD_TYPE = property(_get_CARD_TYPE, _set_CARD_TYPE)
    CARD_NAME  = property(_get_CARD_NAME , _set_CARD_NAME )
    HP = property(_get_HP, _set_HP)
    POKE_TYPE  = property(_get_POKE_TYPE , _set_POKE_TYPE )
    SKILL_COST  = property(_get_SKILL_COST , _set_SKILL_COST )
    SKILL_NAME  = property(_get_SKILL_NAME , _set_SKILL_NAME )
    SKILL_TEXT  = property(_get_SKILL_TEXT , _set_SKILL_TEXT )
    WEAK_TYPE  = property(_get_WEAK_TYPE , _set_WEAK_TYPE )
    WEAK_MULTI = property(_get_WEAK_MULTI, _set_WEAK_MULTI)
    REGI_TYPE  = property(_get_REGI_TYPE , _set_REGI_TYPE )
    RUN_COUNT   = property(_get_RUN_COUNT  , _set_RUN_COUNT  )
    FLAVOR = property(_get_FLAVOR, _set_FLAVOR)
    ILLUSTRATOR = property(_get_ILLUSTRATOR, _set_ILLUSTRATOR)
    SKILL_DAMAGE = property(_get_SKILL_DAMAGE, _set_SKILL_DAMAGE) 
    FROM = property(_get_FROM, _set_FROM)
    POKE_NO = property(_get_POKE_NO, _set_POKE_NO)
    CATEGORY = property(_get_CATEGORY, _set_CATEGORY)	
    HEIGHT = property(_get_HEIGHT, _set_HEIGHT)
    WEIGHT = property(_get_WEIGHT, _set_WEIGHT)
    RARITY = property(_get_RARITY, _set_RARITY)
    REGI_MULTI = property(_get_REGI_MULTI, _set_REGI_MULTI)
    SERIES_MARK = property(_get_SERIES_MARK, _set_SERIES_MARK) 
    REGULATION_MARK = property(_get_REGULATION_MARK, _set_REGULATION_MARK) 
    
    def __str__(self):
        temp = ""
        temp += "-----------------------------------------------------------------------\n"
        
        temp += self._CARD_TYPE[0] #10
        temp += '\t'
        temp += self._CARD_NAME[0]  #14
        
        if((self._CARD_TYPE[0]!="Item")and(self._CARD_TYPE[0]!="Supporter")and(self._CARD_TYPE[0]!="Stadium")and(self._CARD_TYPE[0]!="Special Energy")):#Item과 Suppandter는 HP, self._POKE_TYPE을 입력하지않음
            #temp += "bbbbbbbbbbbbbbbbbbbbbbbbbbb"
            temp += '\t'
            temp += "HP"
            temp += str(self._HP[0]) #16
            temp += '\t'
            temp += self._POKE_TYPE[0]#18
            temp += "\n"
        #print(str(self._FROM[0]))
        #print("check")
        
        if(str(self._FROM[0]) != ""): #From 값이 있을때만 입력
            temp += str(self._FROM[0]) #12
            temp += "에서 진화"
        temp += "\n"

        #GX , TAG, Item, Suppoter, Stadium, Special Energy는 기본정보 출력하지 않음
        
        if((self._CARD_TYPE[0][-2:]!="GX") and (self._CARD_TYPE[0][-3:]!="TAG") and (self._CARD_TYPE[0]!="Item")and(self._CARD_TYPE[0]!="Supporter")and(self._CARD_TYPE[0]!="Stadium")and(self._CARD_TYPE[0]!="Special Energy")): #번호,키,몸무게는 GX가아닌경우만 출력
            temp += "전국도감 NO."
            temp += str(self._POKE_NO[0]) #35
            temp += "  "
            temp += str(self._CATEGORY[0]) #37
            temp += "  키: "
            temp += str(self._HEIGHT[0]).replace("m", " m") #39 #원본 m에서 공백 추가
            temp += "  몸무게: "
            temp += str(self._WEIGHT[0]).replace("kg", " kg")#41
        temp += '\n'

        for i in range(len(self._SKILL_NAME)):
            #print (("str(self._SKILL_COST[%d] : %s") % (i ,str(self._SKILL_COST[i])))
            if(str(self._SKILL_COST[i]) != ""): # attack cost가 비어있지 않을때만
                temp += str(self._SKILL_COST[i]) #30
                temp += '\t'
            if(self._SKILL_NAME[i] != "n/a"):
                if(self._SKILL_NAME[i][:9] == "[Ability]"):
                    temp += "{AB}\t"
                    temp += self._SKILL_NAME[i][10:] #10번재부터 공백한칸제거
                else:
                    temp += self._SKILL_NAME[i] #27
                temp += '\t'
            if(str(self._SKILL_DAMAGE[i]) != "n/a"):
                if( (str(self._SKILL_DAMAGE[i])[-1:] == "+") or (str(self._SKILL_DAMAGE[i])[-1:] == "-") or (str(self._SKILL_DAMAGE[i])[-1:] == "×") ):
                    temp += str(self._SKILL_DAMAGE[i]).replace(str(self._SKILL_DAMAGE[i])[-1:], "\t"+str(self._SKILL_DAMAGE[i])[-1:])
                else:
                    temp += str(self._SKILL_DAMAGE[i]) #30
            
            temp += '\n' #attack text출력전 줄바꿈필요
            if(self._SKILL_TEXT[i] != "n/a"):
                temp += self._SKILL_TEXT[i] #33
                temp += '\n'
            
        if((self._CARD_TYPE[0]!="Item")and(self._CARD_TYPE[0]!="Supporter")and(self._CARD_TYPE[0]!="Stadium")and(self._CARD_TYPE[0]!="Special Energy")): #Item Suppandter는 약점 관련 정보필요없음
            temp += "\n\t"
            temp += self._WEAK_TYPE[0]#20
            temp += self._WEAK_MULTI[0]#21
            temp += '\t'
            temp += self._REGI_TYPE[0]#23 
            #print("regi type")
            #print(str(self._REGI_TYPE[0]))
            temp += str(self._REGI_MULTI[0])#24
            #print("regi multi")
            #print(str(self._REGI_MULTI[0]))
            temp += '\t'
            if(self._RUN_COUNT[0] != "n/a"):
                cnt = int(self._RUN_COUNT[0])#25
                for i in range(cnt):
                    temp += "{C}"
            temp += '\n'

        temp += "Illus. "
        temp += self._ILLUSTRATOR[0] #4
        temp += "\n\n"
        temp += self._CARD_NO[0] #1
        temp += '\t'
        temp += self._RARITY[0] #3
        temp += "\n\n"
        
        if((self._CARD_TYPE[0][-2:]!="GX")and(self._CARD_TYPE[0]!="Item")and(self._CARD_TYPE[0]!="Supporter")and(self._CARD_TYPE[0]!="Stadium")and(self._CARD_TYPE[0]!="Special Energy")): #Item Suppandter는 Flavand Text필요없음
        
            temp += str(self._FLAVOR[0])#42
            temp += '\n'
        
        temp += "-----------------------------------------------------------------------\n"
        return temp   

class KR_DATA:
    def __init__(self, f_name):
        self.file_name = f_name
        #self.writer = pd.ExcelWriter('pandas_xlsxWriter.xlsx', engine='openpyxl', mode='a')
        self.format = {}
    
    def Card_Type_Number(self, type):
        if(type == "Item"):
            return 21
        elif(type == "Supporter"):
            return 22
        elif(type == "Stadium"):
            return 23
        elif(type == "Energy" ):
            return 31
        elif(type == "Special Energy"):
            return 32
        elif(type == "Basic"):
            return 11 #포켓몬
        elif(type == "Stage 1"):
            return 12
        else:
            return 13

    def CardType_Eng_to_Kor(self, ENG):
        ENG = ENG.replace("Basic","기본 포켓몬")
        ENG = ENG.replace("Stage 1","1진화 포켓몬")
        ENG = ENG.replace("Stage 2","2진화 포켓몬")
        ENG = ENG.replace("Stage V","V진화")
        ENG = ENG.replace("/GX/TAG",", TAG TEAM GX")
        ENG = ENG.replace("/GX",", 포켓몬 GX")
        ENG = ENG.replace("/VMAX",", 포켓몬 VMAX") #0220 s1대응
        ENG = ENG.replace("/V",", 포켓몬 V") #0220 s1대응
        
        #기타
        ENG = ENG.replace("/◇Prismstar", ", 프리즘스타")

        #아이템류
        ENG = ENG.replace("Item","아이템")
        ENG = ENG.replace("Supporter","서포트")
        ENG = ENG.replace("Stadium","스타디움")
        ENG = ENG.replace("/EX",", 포켓몬 EX") #??? Sample필요
        
        #에너지
        ENG = ENG.replace("Special Energy","특수 에너지")
        ENG = ENG.replace("Energy","기본 에너지")

        #복원
    
        #포켓몬의 도구
        return ENG
    def Properties_Eng_to_Kor(self, ENG):    
        #{G}풀,{R}불꽃,{W}물,{L}번개,{P}초,{F}격투,{D}악,{M}강철,{S}페어리,{C}무색,{Y}드래곤
        #강철
        ENG = ENG.replace("{M}","강철");
        #강철,격투
        #격투
        ENG = ENG.replace("{F}","격투")
        #드래곤
        ENG = ENG.replace("{Y}","드래곤")
        #무색
        #?? 무색이무속성?
        ENG = ENG.replace("{C}","무색")
        #물
        ENG = ENG.replace("{W}","물")
        #번개
        ENG = ENG.replace("{L}","번개")
        #번개,풀
        #불꽃
        ENG = ENG.replace("{R}","불꽃")
        #불꽃,물
        #악
        ENG = ENG.replace("{D}","악")
        #악,강철
        #초
        ENG = ENG.replace("{P}","초")
        #페어리
        ENG = ENG.replace("{S}","페어리")
        #페어리,물
        #페어리,초
        #풀
        ENG = ENG.replace("{G}","풀")
        #풀,불꽃
        #풀,악
        ENG = ENG.replace("{X}","0코스트") #0220 s1대응
        
        return ENG
    
    def SpecialTextCheck(self, text):
        text = text.replace("[자신]","자신")
        text = text.replace("[상대]","상대")
        text = text.replace("[독]","독")
        text = text.replace("[화상]","화상")
        text = text.replace("[마비]","마비")
        text = text.replace("[잠듦]","잠듦")
        text = text.replace("[혼란]","혼란")
        text = text.replace("[기절]","기절")
        text = text.replace("[기본]","기본")
        text = text.replace("[1진화]","1진화")
        text = text.replace("[2진화]","2진화")
        text = text.replace("[후퇴]","후퇴")
        text = text.replace("[포켓몬 GX 룰]","포켓몬 GX 룰")
        text = text.replace("[TAG TEAM 룰]","TAG TEAM 룰")
        text = text.replace("[아이템 룰]","아이템 룰")
        text = text.replace("[서포트 룰]","서포트 룰")
        text = text.replace("[스타디움 룰]","스타디움 룰")
        text = text.replace("[포켓몬 EX 룰]","포켓몬 EX 룰")
        text = text.replace("[M진화(메가진화) 룰]","M진화 룰")
        text = text.replace("[포켓몬의 도구 룰]","포켓몬의 도구 룰")
        text = text.replace("[Ability]","[특성]")
        text = text.replace("n/a","")

        return text
    
    def BoldTextCheck(self,text):
        text = text.replace("[기본]","[i]기본[/i]")
        text = text.replace("[1진화]","[i]1진화[/i]")
        text = text.replace("[2진화]","[i]2진화[/i]")
        text = text.replace("[기절]","[i]기절[/i]")
        text = text.replace("[후퇴]","[i]후퇴[/i]")
        text = text.replace("[독]","[i]독[/i]")
        text = text.replace("[화상]","[i]화상[/i]")
        text = text.replace("[마비]","[i]마비[/i]")
        text = text.replace("[잠듦]","[i]잠듦[/i]")
        text = text.replace("[혼란]","[i]혼란[/i]") 
        text = text.replace("[VMAX 룰]","[i]VMAX 룰[/i]") #0220 s1대응
        text = text.replace("[VMAX]","[i]VMAX[/i]") #0220 s1대응
        text = text.replace("[V 룰]","[i]V 룰[/i]") #0220 s1대응
        text = text.replace("[V]","[i]V[/i]") #0220 s1대응
        text = text.replace("[GX]","[i]GX[/i]") #0220 s1대응
        text = text.replace("{G}","[s]풀[/s]")
        text = text.replace("{R}","[s]불꽃[/s]")
        text = text.replace("{W}","[s]물[/s]")
        text = text.replace("{L}","[s]번개[/s]")
        text = text.replace("{P}","[s]초[/s]")
        text = text.replace("{C}","[s]무색[/s]")
        text = text.replace("{F}","[s]격투[/s]")
        text = text.replace("{D}","[s]악[/s]")
        text = text.replace("{M}","[s]강철[/s]")
        text = text.replace("{Y}","[s]페어리[/s]")
        text = text.replace("{N}","[s]드래곤[/s]")
        
        return text
        
    def Remove_null_Data (self, token, method, null_val = ""):
        tmp = self.format[token][0]
        if("n/a" in str(tmp)): 
            self.format[token]=[null_val]
        else:
            if(method == int):
                if(str(tmp).find('₋')>=0): #JP Minus '₋' replace to KR '-' for int() funtion
                    tmp = str(tmp).replace('₋','-')
            self.format[token]=[method(tmp)]
        
    
    def InputCardData(self, card):
        
        self.format = {}
        self.format = {
        "WR_COLLECT_NO" : [card.CARD_NO[0]] ,  
        "WR_POKE_NO" : [card.POKE_NO[0]],
        "WR_RARITY" : [card.RARITY[0]],
        "WR_CARDTYPE_NUM" : [""], #Function item support poke stage 에 따라 2자리 숫자
        "WR_CARDTYPE" : [card.CARD_TYPE[0]],
        "WR_NAME" : [card.CARD_NAME[0]],
        "WR_MON_TYPE" : [card.CATEGORY[0]], 
        "WR_HEIGHT" : [card.HEIGHT[0]], 
        "WR_WEIGHT" : [card.WEIGHT[0]], 
        "WR_DOX_INFO" : [card.FLAVOR[0]], 
        "WR_PRE_MON" : [card.FROM[0]],
        "WR_LEVEL" : [""],
        "WR_ILLUSTRATOR" : [card.ILLUSTRATOR[0]],
        "WR_HP" : [card.HP[0]],
        "WR_CARDMONTYPE" : [card.POKE_TYPE[0]],
        "WR_WEAKNESS" : [card.WEAK_TYPE[0]],
        "WR_WEAKNESS_LVL" : [card.WEAK_MULTI[0]],
        "WR_RESISTANCE" : [card.REGI_TYPE[0]],
        "WR_RESISTANCE_LVL" : [card.REGI_MULTI[0]],
        "WR_RETREAT" : [card.RUN_COUNT[0]],    
        

        "WR_TECHERG1" : [""],
        "WR_TECHERG2" : [""],
        "WR_TECHERG3" : [""],
        "WR_TECHERG4" : [""],

        "WR_DAMAGE1" : [""],
        "WR_DAMAGE2" : [""],
        "WR_DAMAGE3" : [""],
        "WR_DAMAGE4" : [""],
        
        "WR_AB_LABEL1" :[""],
        "WR_AB_LABEL2" :[""],
        "WR_AB_LABEL3" :[""],
        "WR_AB_LABEL4" :[""],

        "WR_ABILITY1" :[""],
        "WR_ABILITY2" :[""],
        "WR_ABILITY3" :[""],
        "WR_ABILITY4" :[""],

        "WR_POWERBODY" : [""],
        "WR_LEVEL_RULE" : [""],
        
        "WR_GOODS_NAME" : [""],
        "WR_GOODS_NAME2" : [""],
        "WR_GOODS_NAME3" : [""],
        "WR_GOODS_NAME4" : [""],
        "WR_GOODS_NAME_SHORT" : [""],
        "WR_GOODS_NAME_IMAGE" : [card.SERIES_MARK[0]],
        "WR_CARDTYPE_ORDER" : [""],
        "WR_RAR_LVL_ORDER" : [""],
        "WR_REGULATION_MARK" :  [card.REGULATION_MARK[0]]
        }

        
        for i in range(4):
            if( i < len(card.SKILL_NAME) ):
                self.format["WR_TECHERG"+str(i+1)] = self.Properties_Eng_to_Kor(card.SKILL_COST[i]).replace("{+}","플러스")
                
                if( str(type(card.SKILL_DAMAGE[i])) == "<class 'int'>" ):
                    self.format["WR_DAMAGE"+str(i+1)] = card.SKILL_DAMAGE[i]
                else:
                    self.format["WR_DAMAGE"+str(i+1)] = card.SKILL_DAMAGE[i].replace("n/a","")
                    
                self.format["WR_AB_LABEL"+str(i+1)] = self.BoldTextCheck(self.SpecialTextCheck(card.SKILL_NAME[i]))
                self.format["WR_ABILITY"+str(i+1)] = self.BoldTextCheck(card.SKILL_TEXT[i].replace("n/a",""))
        
        
        
        self.format["WR_CARDTYPE_NUM"] = self.Card_Type_Number(self.format["WR_CARDTYPE"][0])
        self.format["WR_CARDTYPE"] = self.CardType_Eng_to_Kor(self.format["WR_CARDTYPE"][0])  
        
        index = self.format["WR_NAME"][0].find("[울트라비스트]")
        if( index >= 0 ):#울트라 비스트인경우
            self.format["WR_NAME"][0] = [ self.format["WR_NAME"][0][0:index-1] ]
        
        #210127 신규Label대응 [일격] , [다이맥스/일격], [거다이맥스/일격], [아이템/일격], [포켓몬의 도구/일격], [서포트/일격], [스타디움/일격], [특수 에너지/일격]
        #210127 신규Label대응 [연격] , [다이맥스/연격], [거다이맥스/연격], [아이템/연격], [포켓몬의 도구/연격], [서포트/연격], [스타디움/연격], [특수 에너지/연격]
        index=self.format["WR_NAME"][0].find("격]")         
        if( index >= 0 ):#일격,연격 카드인경우
            index = self.format["WR_NAME"][0].find("[")
            if( self.format["WR_NAME"][0].find("연격]") > 0):
                self.format["WR_NAME"][0] = self.format["WR_NAME"][0][0:index-1]+"[yk]연격[/yk]"  #210221 CardType ,연격추가 및 yk변경
                self.format["WR_CARDTYPE"] = self.format["WR_CARDTYPE"] + ",연격"
            else:
                self.format["WR_NAME"][0] = self.format["WR_NAME"][0][0:index-1]+"[ik]일격[/ik]"  #210221 CardType ,일격추가 및 ik변경
                self.format["WR_CARDTYPE"] = self.format["WR_CARDTYPE"] + ",일격"
        
        self.format["WR_NAME"][0] = self.format["WR_NAME"][0].replace("[다이맥스]","") #210127 불필요 Label제거
        self.format["WR_NAME"][0] = self.format["WR_NAME"][0].replace("[거다이맥스]","") #210127 불필요 Label제거

        self.format["WR_NAME"][0] = self.Properties_Eng_to_Kor(self.format["WR_NAME"][0]) #0905 CardName 영문 한글변환 적용
            
        
        self.Remove_null_Data("WR_POKE_NO" , int , 0)

        self.format["WR_RARITY"]=self.format["WR_RARITY"][0][1:-1] #remove ( )

        
        if(self.format["WR_MON_TYPE"][0] == 0): #GX인경우 Data가 0으로 들어가있으며 공백으로 전환필요
            self.format["WR_MON_TYPE"] = ""

        
        self.format["WR_HEIGHT"] = self.format["WR_HEIGHT"][0].replace("m", "")
        if( self.format["WR_HEIGHT"] == ""): #GX인경우 0으로 입력
            self.format["WR_HEIGHT"] = 0
        elif (  self.format["WR_HEIGHT"].find(".") > 0):
            self.format["WR_HEIGHT"] = float(self.format["WR_HEIGHT"])
        else:
            self.format["WR_HEIGHT"] = int(self.format["WR_HEIGHT"])
        
        self.format["WR_WEIGHT"] = self.format["WR_WEIGHT"][0].replace("kg", "")
        if( self.format["WR_WEIGHT"] == ""): #GX인경우 0으로 입력
            self.format["WR_WEIGHT"] = 0
        elif (  self.format["WR_WEIGHT"].find(".") > 0):
            self.format["WR_WEIGHT"] = float(self.format["WR_WEIGHT"])
        else:
            self.format["WR_WEIGHT"] = int(self.format["WR_WEIGHT"])

            
        self.format["WR_DOX_INFO"] = self.format["WR_DOX_INFO"][0].replace("\n"," ")

        self.format["WR_ILLUSTRATOR"] = self.format["WR_ILLUSTRATOR"][0].replace("n/a","")        
        
        self.Remove_null_Data("WR_HP", int, null_val=0) #200221 Null Value 0으로 변경
        
        self.format["WR_CARDMONTYPE"] = self.Properties_Eng_to_Kor(self.format["WR_CARDMONTYPE"][0]) #한글변경
        self.format["WR_WEAKNESS"][0] = self.Properties_Eng_to_Kor(self.format["WR_WEAKNESS"][0])
        
        self.Remove_null_Data("WR_WEAKNESS_LVL",str)
 
        self.format["WR_RESISTANCE"] = self.Properties_Eng_to_Kor(self.format["WR_RESISTANCE"][0])
        
        self.Remove_null_Data("WR_RESISTANCE_LVL",int)
        
        self.Remove_null_Data("WR_RETREAT",int, null_val=0)
        
        #Make Data frame , Append to Excel File
        #df = pd.DataFrame( self.format )
        df = DataFrame( self.format )

        return df
        #self.append_df_to_excel(df , startcol = 2, index = False,  header = False)

        #df.to_excel(self.writer, sheet_name='Sheet1', startrow = 0, startcol = 2, index = False,  header = False)
        #df.to_excel(self.writer, sheet_name='Sheet1')
        
        
    def close_writer(self):
        #self.writer.close()
        return 0
        
    
 

class ConverterThread(QThread):
    
    isFinish = pyqtSignal(list)

    def __init__(self):
        super().__init__()
        self.finish = False
        self.MsgSlot = None
        self.r_fileName = None
        self.r_data = 0 #read data (JP Excel file)
        self.w_data = 0 #write data (KR Excel file)


    @pyqtSlot(list)
    def ConvertMsgSlot(self, Msg):
        self.MsgSlot = Msg[0]
        if(len(Msg) > 1): #Load Case
            self.r_fileName = Msg[1]
        

    def run(self):
        while True:
            #print("[ConverterThread] Wait MSG")
            if(self.MsgSlot == "LOAD"):
                print("[ConverterThread] LOAD!!")
                self.Load_JP()
                if(self.r_data.isLoadOK):
                    self.isFinish.emit([True, self.r_data]) #Read Fail case self.r_data == None
                else:
                    self.isFinish.emit([True, None]) #Read Fail casee
                self.MsgSlot = None
            elif(self.MsgSlot == "CVT_KR"):
                print("[ConverterThread] CVT_KR")
                self.Convert_KR()
                self.MsgSlot = None
                self.isFinish.emit([True])
            elif(self.MsgSlot == "CVT_TXT"):
                print("[ConverterThread] CVB_TXT")
                self.Convert_Txt()
                self.MsgSlot = None
                self.isFinish.emit([True])
            elif(self.MsgSlot == "FINISH"):
                self.MsgSlot = None
                self.isFinish.emit([True])
                break
            self.msleep(500)
            
        print("[ConverterThread] Thread Finish")
        
    
    def Load_JP(self):
        print ("[ConverterThread] LOAD_JP")
        print(self.r_fileName)
        self.r_data = JP_DATA(self.r_fileName)

    
    def Convert_KR(self):
        print ("[ConverterThread] CONVERT_KR")
        write_name = os.path.splitext(self.r_fileName)[0]+"_KOR.xlsx"
        self.w_data = KR_DATA(write_name)
        self.file_name = write_name
        df = DataFrame()
        while(True):
            card = self.r_data.Get_Card_Data()
            if(card == "EOF"):
                break
            print("NOW_CARD : %s" % card.CARD_NO[0])
            df = concat([df, self.w_data.InputCardData(card)])
        
        self.append_df_to_excel(df , startcol = 2, index = False,  header = False)
    
    def Convert_Txt(self):
        print ("[ConverterThread] CONVERT_TXT")
        write_name = os.path.splitext(self.r_fileName)[0]+".txt"
        self.r_data.SetStartPos() #ready to convert again
        self.r_data.export_txt_files(write_name)

    def closeEvent(self, event):
        print("[ConverterThread] closeEvent!!!")
        self.terminate()
    
    def append_df_to_excel(self, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
        '''
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          df : dataframe to save to workbook
          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        '''
        from openpyxl import load_workbook

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        #writer = pd.ExcelWriter(self.file_name, engine='openpyxl')
        writer = ExcelWriter(self.file_name, engine='openpyxl')

        try:
            # try to open an existing workbook
            writer.book = load_workbook(self.file_name)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0
        
        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()



